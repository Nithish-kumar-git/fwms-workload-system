"""
Reports service — workload reporting and export generation.
Spec reference: final_system_specification.md (Reporting System)

Provides:
  1. Faculty workload report (per-faculty with subject details)
  2. Subject-wise summary (per-offering with assigned faculty)
  3. Department summary (aggregate statistics)
  4. Excel export (openpyxl — 3 sheets)
  5. PDF export (basic HTML-to-text table)

All read-only queries, no mutations.
"""

from __future__ import annotations
from typing import Optional
from sqlalchemy import text
from app.db.session import get_transaction
import logging
import io
from datetime import datetime

logger = logging.getLogger(__name__)


def _resolve_active_cycle(session) -> tuple[str, int]:
    """
    Resolve academic_year and semester_id from the active cycle.

    Returns:
        (academic_year, semester_id) from cycle table with status = 'OPEN'

    Raises:
        RuntimeError: if no active cycle exists
    """
    row = session.execute(
        text("""
            SELECT ay.name, c.semester_id
            FROM cycle c
            JOIN academic_year ay ON ay.id = c.academic_year_id
            WHERE c.status = 'OPEN'
            LIMIT 1
        """)
    ).fetchone()

    if not row:
        raise RuntimeError("No active cycle found. Activate a cycle before generating reports.")

    return row[0], row[1]


# ============================================================================
# STEP 1: Faculty Workload Report
# ============================================================================

def get_faculty_workload(
    academic_year: Optional[str] = None, semester_id: Optional[int] = None
) -> dict:
    """Per-faculty workload report with assigned subjects."""
    with get_transaction() as session:
        if academic_year is None or semester_id is None:
            academic_year, semester_id = _resolve_active_cycle(session)
        # Get all faculty with their norms
        faculty_rows = session.execute(
            text("""
                SELECT s.id, s.emp_code, s.name, s.designation,
                       COALESCE(s.tch_norm, 40) AS tch_norm
                FROM staff s
                WHERE s.emp_code IS NOT NULL AND s.is_active = true
                ORDER BY s.designation, s.name
            """)
        ).fetchall()

        records = []
        for f in faculty_rows:
            staff_id = f[0]

            # Get subjects assigned to this faculty
            subj_rows = session.execute(
                text("""
                    SELECT sub.code, sub.name, p.name AS program,
                           sem.label AS semester, sec.label AS section,
                           COALESCE(sub.l, 0), COALESCE(sub.t, 0),
                           COALESCE(sub.p, 0), COALESCE(sub.tch, 0)
                    FROM allocation a
                    JOIN subject_offering so ON so.id = a.subject_offering_id
                    JOIN subject sub ON sub.id = so.subject_id
                    JOIN program p ON p.id = so.program_id
                    JOIN semester sem ON sem.id = so.semester_id
                    JOIN section sec ON sec.id = so.section_id
                    WHERE a.staff_id = :sid
                      AND so.academic_year = :year
                      AND so.semester_id = :sem_id
                    ORDER BY p.name, sem.label, sec.label
                """),
                {"sid": staff_id, "year": academic_year, "sem_id": semester_id}
            ).fetchall()

            subjects = [
                {
                    "course_code": r[0], "course_name": r[1],
                    "program": r[2], "semester": r[3], "section": r[4],
                    "l": r[5], "t": r[6], "p": r[7], "tch": r[8],
                }
                for r in subj_rows
            ]

            assigned_tch = sum(s["tch"] for s in subjects)
            tch_norm = f[4]

            records.append({
                "staff_id": staff_id, "emp_code": f[1], "name": f[2],
                "designation": f[3], "assigned_tch": assigned_tch,
                "tch_norm": tch_norm,
                "deviation_hours": assigned_tch - tch_norm,
                "subjects_assigned": subjects,
            })

    return {"total_faculty": len(records), "records": records}


# ============================================================================
# STEP 2: Subject-Wise Summary
# ============================================================================

def get_subject_summary(
    academic_year: Optional[str] = None, semester_id: Optional[int] = None
) -> dict:
    """Per-subject-offering report showing assigned faculty."""
    with get_transaction() as session:
        if academic_year is None or semester_id is None:
            academic_year, semester_id = _resolve_active_cycle(session)
        rows = session.execute(
            text("""
                SELECT so.id, sub.code, sub.name, p.name AS program,
                       sem.label AS semester, sec.label AS section,
                       s.name AS faculty_name, s.emp_code,
                       COALESCE(sub.tch, 0) AS tch,
                       CASE WHEN a.id IS NOT NULL THEN true ELSE false END AS allocated
                FROM subject_offering so
                JOIN subject sub ON sub.id = so.subject_id
                JOIN program p ON p.id = so.program_id
                JOIN semester sem ON sem.id = so.semester_id
                JOIN section sec ON sec.id = so.section_id
                LEFT JOIN allocation a ON a.subject_offering_id = so.id
                LEFT JOIN staff s ON s.id = a.staff_id
                WHERE so.academic_year = :year AND so.semester_id = :sem_id
                ORDER BY p.name, sem.label, sec.label, sub.code
            """),
            {"year": academic_year, "sem_id": semester_id}
        ).fetchall()

    records = [
        {
            "subject_offering_id": r[0],
            "course_code": r[1], "course_name": r[2], "program": r[3],
            "semester": r[4], "section": r[5],
            "faculty_name": r[6], "faculty_emp_code": r[7],
            "tch": r[8], "allocated": r[9],
        }
        for r in rows
    ]
    return {"total": len(records), "records": records}


# ============================================================================
# STEP 3: Department Summary
# ============================================================================

def get_department_summary(
    academic_year: Optional[str] = None, semester_id: Optional[int] = None
) -> dict:
    """Aggregate department statistics."""
    with get_transaction() as session:
        if academic_year is None or semester_id is None:
            try:
                academic_year, semester_id = _resolve_active_cycle(session)
            except RuntimeError:
                # No active cycle - return safe empty response
                total_faculty = session.execute(
                    text("SELECT count(*) FROM staff WHERE emp_code IS NOT NULL AND is_active = true")
                ).scalar()
                return {
                    "total_subject_offerings": 0,
                    "allocated_subjects": 0,
                    "unallocated_subjects": 0,
                    "total_faculty": total_faculty or 0,
                    "average_workload": 0.0,
                    "faculty_overloaded": 0,
                    "faculty_underloaded": 0,
                    "faculty_balanced": total_faculty or 0,
                }
        
        total_offerings = session.execute(
            text("""
                SELECT count(*) FROM subject_offering
                WHERE academic_year = :year AND semester_id = :sem_id
                  AND is_active = true
            """),
            {"year": academic_year, "sem_id": semester_id}
        ).scalar()

        allocated = session.execute(
            text("""
                SELECT count(DISTINCT a.subject_offering_id)
                FROM allocation a
                JOIN subject_offering so ON so.id = a.subject_offering_id
                WHERE so.academic_year = :year AND so.semester_id = :sem_id
            """),
            {"year": academic_year, "sem_id": semester_id}
        ).scalar()

        total_faculty = session.execute(
            text("SELECT count(*) FROM staff WHERE emp_code IS NOT NULL AND is_active = true")
        ).scalar()

        # Convert semester_id to semester_type for workload_summary table (legacy schema)
        semester_type = "EVEN" if semester_id in (2, 4, 6) else "ODD"
        
        avg_workload = session.execute(
            text("""
                SELECT COALESCE(AVG(ws.tch_total), 0)
                FROM workload_summary ws
                WHERE ws.academic_year = :year AND ws.semester_type = :sem_type
            """),
            {"year": academic_year, "sem_type": semester_type}
        ).scalar()

        overloaded = session.execute(
            text("""
                SELECT count(*) FROM workload_summary
                WHERE academic_year = :year AND semester_type = :sem_type
                  AND deviation_hours > 0
            """),
            {"year": academic_year, "sem_type": semester_type}
        ).scalar()

        underloaded = session.execute(
            text("""
                SELECT count(*) FROM workload_summary
                WHERE academic_year = :year AND semester_type = :sem_type
                  AND deviation_hours < -2
            """),
            {"year": academic_year, "sem_type": semester_type}
        ).scalar()

    unallocated = total_offerings - allocated
    balanced = total_faculty - overloaded - underloaded

    return {
        "total_subject_offerings": total_offerings,
        "allocated_subjects": allocated,
        "unallocated_subjects": unallocated,
        "total_faculty": total_faculty,
        "average_workload": round(float(avg_workload), 2),
        "faculty_overloaded": overloaded,
        "faculty_underloaded": underloaded,
        "faculty_balanced": max(balanced, 0),
    }


# ============================================================================
# STEP 4: Excel Export
# ============================================================================

def generate_excel_report(
    academic_year: Optional[str] = None, semester_id: Optional[int] = None
) -> bytes:
    """Generate Excel workbook with 3 sheets. Returns bytes. Never raises."""
    # Resolve active cycle if not provided — graceful fallback
    no_cycle = False
    if academic_year is None or semester_id is None:
        try:
            with get_transaction() as session:
                academic_year, semester_id = _resolve_active_cycle(session)
        except Exception as e:
            logger.warning(f"No active cycle: {e}")
            academic_year = "N/A"
            semester_id = 0
            no_cycle = True

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # --- Style definitions ---
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    warn_font = Font(bold=True, size=11, color="CC0000")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    data_font = Font(size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    def style_data(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

    def auto_size_columns(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    def add_note_row(ws):
        """Insert a warning row if no active cycle was found."""
        if no_cycle:
            ws.append(["No active academic cycle found — showing empty report"])
            ws.cell(row=ws.max_row, column=1).font = warn_font

    # ---- Fetch data (safe — empty dicts if query fails) ----
    try:
        faculty_data = get_faculty_workload(academic_year, semester_id)
    except Exception as e:
        logger.warning(f"Faculty workload query failed: {e}")
        faculty_data = {"total_faculty": 0, "records": []}

    try:
        subj_data = get_subject_summary(academic_year, semester_id)
    except Exception as e:
        logger.warning(f"Subject summary query failed: {e}")
        subj_data = {"total": 0, "records": []}


    # Row offset: if no_cycle, row 1 = warning, row 2 = header
    header_row = 2 if no_cycle else 1
    data_start = header_row + 1

    # ---- Sheet 1: Faculty Workload ----
    ws1 = wb.active
    ws1.title = "Faculty Workload"
    add_note_row(ws1)
    ws1.append([
        "Emp Code", "Name", "Designation", "TCH Norm",
        "Assigned TCH", "Deviation", "Course Code", "Course Name",
        "Program", "Semester", "Section", "L", "T", "P", "TCH",
    ])
    style_header(ws1, row=header_row)

    if faculty_data["records"]:
        for fac in faculty_data["records"]:
            if fac["subjects_assigned"]:
                for i, subj in enumerate(fac["subjects_assigned"]):
                    ws1.append([
                        fac["emp_code"] if i == 0 else "",
                        fac["name"] if i == 0 else "",
                        fac["designation"] if i == 0 else "",
                        fac["tch_norm"] if i == 0 else "",
                        fac["assigned_tch"] if i == 0 else "",
                        fac["deviation_hours"] if i == 0 else "",
                        subj["course_code"], subj["course_name"],
                        subj["program"], subj["semester"], subj["section"],
                        subj["l"], subj["t"], subj["p"], subj["tch"],
                    ])
            else:
                ws1.append([
                    fac["emp_code"], fac["name"], fac["designation"],
                    fac["tch_norm"], 0, -fac["tch_norm"],
                    "", "", "", "", "", "", "", "", "",
                ])

    auto_size_columns(ws1)
    style_data(ws1, start_row=data_start)

    # ---- Sheet 2: Subject Summary ----
    ws2 = wb.create_sheet("Subject Summary")
    add_note_row(ws2)
    ws2.append([
        "Course Code", "Course Name", "Program", "Semester",
        "Section", "Faculty", "Emp Code", "TCH", "Allocated",
    ])
    style_header(ws2, row=header_row)

    if subj_data["records"]:
        for rec in subj_data["records"]:
            ws2.append([
                rec["course_code"], rec["course_name"], rec["program"],
                rec["semester"], rec["section"],
                rec["faculty_name"] or "UNASSIGNED",
                rec["faculty_emp_code"] or "",
                rec["tch"], "Yes" if rec["allocated"] else "No",
            ])

    auto_size_columns(ws2)
    style_data(ws2, start_row=data_start)

    # ---- Sheet 3: Workload Summary ----
    ws3 = wb.create_sheet("Workload Summary")
    add_note_row(ws3)
    ws3.append([
        "Emp Code", "Name", "Designation", "TCH Norm",
        "TCH Assigned", "Deviation", "Status",
    ])
    style_header(ws3, row=header_row)

    if faculty_data["records"]:
        for fac in faculty_data["records"]:
            deviation = fac["deviation_hours"]
            if deviation > 0:
                status = "OVERLOADED"
            elif deviation < -2:
                status = "UNDERLOADED"
            else:
                status = "BALANCED"
            ws3.append([
                fac["emp_code"], fac["name"], fac["designation"],
                fac["tch_norm"], fac["assigned_tch"], deviation, status,
            ])

    auto_size_columns(ws3)
    style_data(ws3, start_row=data_start)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# STEP 5: PDF Export (simple HTML table → text)
# ============================================================================

def generate_pdf_report(
    academic_year: Optional[str] = None, semester_id: Optional[int] = None
) -> bytes:
    """
    Generate a simple PDF report.
    Uses reportlab if available, otherwise falls back to plain text.
    """
    if academic_year is None or semester_id is None:
        with get_transaction() as session:
            academic_year, semester_id = _resolve_active_cycle(session)

    assert academic_year is not None and semester_id is not None

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        return _generate_reportlab_pdf(academic_year, semester_id)
    except ImportError:
        return _generate_text_pdf(academic_year, semester_id)


def _generate_reportlab_pdf(academic_year: str, semester_id: int) -> bytes:
    """Generate PDF using reportlab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(
        f"Faculty Workload Report — {academic_year} (Semester {semester_id})",
        styles["Title"]
    ))
    elements.append(Spacer(1, 20))

    # Faculty Workload Table
    faculty_data = get_faculty_workload(academic_year, semester_id)
    table_data = [["Emp Code", "Name", "Designation", "Norm", "Assigned", "Deviation"]]
    
    normal_style = styles["Normal"]
    normal_style.fontSize = 9

    for fac in faculty_data["records"]:
        table_data.append([
            fac["emp_code"], 
            Paragraph(fac["name"], normal_style), 
            Paragraph(fac["designation"], normal_style),
            str(fac["tch_norm"]), str(fac["assigned_tch"]),
            str(fac["deviation_hours"]),
        ])

    table = Table(table_data, repeatRows=1, colWidths=[60, 160, 160, 50, 50, 50])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)

    # Department Summary
    elements.append(Spacer(1, 30))
    dept = get_department_summary(academic_year, semester_id)
    elements.append(Paragraph("Department Summary", styles["Heading2"]))
    summary_data = [
        ["Total Offerings", str(dept["total_subject_offerings"])],
        ["Allocated", str(dept["allocated_subjects"])],
        ["Unallocated", str(dept["unallocated_subjects"])],
        ["Avg Workload", str(dept["average_workload"])],
        ["Overloaded", str(dept["faculty_overloaded"])],
        ["Underloaded", str(dept["faculty_underloaded"])],
    ]
    st = Table(summary_data, colWidths=[200, 100])
    st.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(st)

    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def _generate_text_pdf(academic_year: str, semester_id: int) -> bytes:
    """Fallback: plain text report when reportlab is not available."""
    faculty_data = get_faculty_workload(academic_year, semester_id)
    dept = get_department_summary(academic_year, semester_id)

    lines = []
    lines.append(f"FACULTY WORKLOAD REPORT — {academic_year} (Semester {semester_id})")
    lines.append("=" * 80)
    lines.append("")
    lines.append(f"{'Emp Code':<10} {'Name':<25} {'Designation':<22} {'Norm':>5} {'Assigned':>9} {'Dev':>5}")
    lines.append("-" * 80)

    for fac in faculty_data["records"]:
        lines.append(
            f"{fac['emp_code']:<10} {fac['name']:<25} {fac['designation']:<22} "
            f"{fac['tch_norm']:>5} {fac['assigned_tch']:>9} {fac['deviation_hours']:>5}"
        )

    lines.append("")
    lines.append("DEPARTMENT SUMMARY")
    lines.append("-" * 40)
    lines.append(f"Total Offerings:  {dept['total_subject_offerings']}")
    lines.append(f"Allocated:        {dept['allocated_subjects']}")
    lines.append(f"Unallocated:      {dept['unallocated_subjects']}")
    lines.append(f"Avg Workload:     {dept['average_workload']}")
    lines.append(f"Overloaded:       {dept['faculty_overloaded']}")
    lines.append(f"Underloaded:      {dept['faculty_underloaded']}")
    lines.append("")
    lines.append(f"Generated: {datetime.now().isoformat()}")

    return "\n".join(lines).encode("utf-8")
