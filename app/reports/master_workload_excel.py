"""
Master Workload Excel Generator — institutional reference sheet reproduction.

Generates a pixel-perfect Excel file matching the department master workload
sheet format with:
  - 4-row institutional header (yellow, bold, merged)
  - 25-column table header (blue #4F81BD, white bold, wrapped)
  - Per-faculty data blocks with vertical merging
  - Manual calculations: L+T+P, TCH, Deviation, Total Workload
  - Fixed column widths, wrap_text on text-heavy columns
  - Faculty sorted by emp_code ASC
  - Semester displayed as Roman numerals

All data sourced from: allocation, subject_offering, subject, program,
semester, section, staff, workload_summary tables.
"""

from __future__ import annotations
from typing import Optional
import io
import logging
from collections import defaultdict

from sqlalchemy import text
from app.db.session import get_transaction

logger = logging.getLogger(__name__)

# ─── Column definitions (1-indexed in Excel, 0-indexed here) ─────────────────

COLUMNS = [
    "S No",                     # A  (0)
    "Emp Code",                 # B  (1)
    "Faculty Name",             # C  (2)
    "Designation",              # D  (3)
    "UG/PG",                    # E  (4)
    "Programme",                # F  (5)
    "Course Category",          # G  (6)
    "SEM",                      # H  (7)
    "Section",                  # I  (8)
    "Student Strength",         # J  (9)
    "COURSE CODE",              # K  (10)
    "COURSE NAME",              # L  (11)
    "Complexity",               # M  (12)
    "CREDITS",                  # N  (13)
    "L",                        # O  (14)
    "T",                        # P  (15)
    "P",                        # Q  (16)
    "L+T+P",                    # R  (17)
    "TCH",                      # S  (18)
    "Minimum Workload",         # T  (19)
    "Deviation",                # U  (20)
    "Remarks",                  # V  (21)
    "Other Academic Engagement",# W  (22)
    "Total Workload",           # X  (23)
    "No. of Research Scholars", # Y  (24)
]

NUM_COLS = len(COLUMNS)  # 25

# Columns that get vertically merged per faculty (0-indexed)
MERGE_COLS = [0, 1, 2, 3, 19, 20, 21, 22, 23, 24]

# Fixed column widths (1-indexed: col_num → width)
COL_WIDTHS = {
    1:  5,    # S No
    2:  10,   # Emp Code
    3:  22,   # Faculty Name
    4:  20,   # Designation
    5:  6,    # UG/PG
    6:  18,   # Programme
    7:  14,   # Course Category
    8:  5,    # SEM
    9:  8,    # Section
    10: 10,   # Student Strength
    11: 14,   # Course Code
    12: 30,   # Course Name
    13: 10,   # Complexity
    14: 8,    # Credits
    15: 4,    # L
    16: 4,    # T
    17: 4,    # P
    18: 7,    # L+T+P
    19: 6,    # TCH
    20: 12,   # Minimum Workload
    21: 10,   # Deviation
    22: 18,   # Remarks
    23: 14,   # Other Academic Engagement
    24: 12,   # Total Workload
    25: 12,   # No. of Research Scholars
}

# Columns that need wrap_text (0-indexed)
WRAP_COLS = {5, 11, 21}  # Programme, Course Name, Remarks

# Semester number → Roman numeral
ROMAN = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI",
          "I": "I", "II": "II", "III": "III", "IV": "IV", "V": "V", "VI": "VI"}


def _to_roman(sem_value: str) -> str:
    """Convert semester label to Roman numeral if it's a digit."""
    s = str(sem_value).strip()
    return ROMAN.get(s, s)


# ─── Data fetching ───────────────────────────────────────────────────────────

def _resolve_active_cycle(session) -> tuple[str, int]:
    """Get active academic_year, semester_id."""
    row = session.execute(
        text("""
            SELECT academic_year, semester_id
            FROM cycle
            WHERE is_active = true
            LIMIT 1
        """)
    ).fetchone()
    if not row:
        raise RuntimeError("No active cycle found.")
    return row[0], row[1]


def _fetch_workload_data(
    academic_year: Optional[str] = None,
    semester_id: Optional[int] = None,
) -> tuple[list[dict], str, int]:
    """
    Fetch all allocation data joined with offering, subject, program,
    semester, section, and staff.  Returns flat list of dicts, one per
    allocation row.  Sorted by emp_code ASC.
    """
    with get_transaction() as session:
        if academic_year is None or semester_id is None:
            academic_year, semester_id = _resolve_active_cycle(session)

        rows = session.execute(
            text("""
                SELECT
                    s.id            AS staff_id,
                    s.emp_code,
                    s.name          AS faculty_name,
                    s.designation,
                    p.ug_pg,
                    p.name          AS programme,
                    COALESCE(sub.course_category, '')  AS course_category,
                    sem.label       AS semester,
                    sec.label       AS section,
                    COALESCE(so.student_strength, 0)   AS student_strength,
                    sub.code        AS course_code,
                    sub.name        AS course_name,
                    COALESCE(a.complexity, '')          AS complexity,
                    COALESCE(sub.credits, 0)           AS credits,
                    COALESCE(a.l_assigned, 0)           AS l_assigned,
                    COALESCE(a.t_assigned, 0)           AS t_assigned,
                    COALESCE(a.p_assigned, 0)           AS p_assigned,
                    COALESCE(ws.norm_hours, 12)         AS norm_hours,
                    COALESCE(ws.other_academic, 0)      AS other_academic,
                    COALESCE(ws.remarks, '')             AS remarks,
                    ws.research_scholars
                FROM allocation a
                JOIN subject_offering so ON so.id = a.subject_offering_id
                JOIN subject sub         ON sub.id = so.subject_id
                JOIN program p           ON p.id = so.program_id
                JOIN semester sem        ON sem.id = so.semester_id
                JOIN section sec         ON sec.id = so.section_id
                JOIN staff s             ON s.id = a.staff_id
                JOIN cycle c             ON c.id = a.cycle_id
                LEFT JOIN workload_summary ws
                    ON ws.staff_id = s.id
                   AND ws.academic_year = :year
                   AND ws.semester_id = :sem_id
                WHERE c.academic_year = :year
                  AND c.semester_id = :sem_id
                  AND s.is_active = true
                ORDER BY s.emp_code ASC, p.name, sem.label, sec.label
            """),
            {"year": academic_year, "sem_id": semester_id},
        ).fetchall()

        data = []
        for r in rows:
            l_val = r[14] or 0
            t_val = r[15] or 0
            p_val = r[16] or 0
            ltp = l_val + t_val + p_val  # FIX #1: compute manually

            data.append({
                "staff_id":         r[0],
                "emp_code":         r[1] or "",
                "faculty_name":     r[2] or "",
                "designation":      r[3] or "",
                "ug_pg":            r[4] or "",
                "programme":        r[5] or "",
                "course_category":  r[6],
                "semester":         r[7] or "",
                "section":          r[8] or "",  # FIX #7: use as stored
                "student_strength": r[9],
                "course_code":      r[10] or "",
                "course_name":      r[11] or "",
                "complexity":       r[12],
                "credits":          r[13],
                "l_assigned":       l_val,
                "t_assigned":       t_val,
                "p_assigned":       p_val,
                "ltp":              ltp,
                "tch":              ltp,  # TCH = L+T+P
                "norm_hours":       r[17],
                "other_academic":   r[18],
                "remarks":          r[19],
                "research_scholars": r[20],  # may be None — DO NOT auto-fill
            })

        # Also fetch faculty who have NO allocations but are active
        unassigned = session.execute(
            text("""
                SELECT s.id, s.emp_code, s.name, s.designation,
                       COALESCE(ws.norm_hours, 12),
                       COALESCE(ws.other_academic, 0),
                       COALESCE(ws.remarks, ''),
                       ws.research_scholars
                FROM staff s
                LEFT JOIN workload_summary ws
                    ON ws.staff_id = s.id
                   AND ws.academic_year = :year
                   AND ws.semester_id = :sem_id
                WHERE s.emp_code IS NOT NULL
                  AND s.is_active = true
                  AND s.id NOT IN (
                      SELECT DISTINCT a2.staff_id
                      FROM allocation a2
                      JOIN cycle c2 ON c2.id = a2.cycle_id
                      WHERE c2.academic_year = :year AND c2.semester_id = :sem_id
                  )
                ORDER BY s.emp_code ASC
            """),
            {"year": academic_year, "sem_id": semester_id},
        ).fetchall()

        for r in unassigned:
            data.append({
                "staff_id":         r[0],
                "emp_code":         r[1] or "",
                "faculty_name":     r[2] or "",
                "designation":      r[3] or "",
                "ug_pg":            "",
                "programme":        "",
                "course_category":  "",
                "semester":         "",
                "section":          "",
                "student_strength": 0,
                "course_code":      "",
                "course_name":      "",
                "complexity":       "",
                "credits":          0,
                "l_assigned":       0,
                "t_assigned":       0,
                "p_assigned":       0,
                "ltp":              0,
                "tch":              0,
                "norm_hours":       r[4],
                "other_academic":   r[5],
                "remarks":          r[6],
                "research_scholars": r[7],
            })

        return data, academic_year, semester_id


def _group_by_faculty(data: list[dict]) -> list[dict]:
    """Group flat rows into per-faculty blocks, sorted by emp_code ASC."""
    grouped = defaultdict(list)
    meta = {}

    for row in data:
        sid = row["staff_id"]
        grouped[sid].append(row)
        if sid not in meta:
            meta[sid] = {
                "emp_code":       row["emp_code"],
                "faculty_name":   row["faculty_name"],
                "designation":    row["designation"],
                "norm_hours":     row["norm_hours"],
                "other_academic": row["other_academic"],
                "remarks":        row["remarks"],
                "research_scholars": row["research_scholars"],
            }

    result = []
    for sid, rows in grouped.items():
        m = meta[sid]
        # FIX #1 & #2: compute totals once per faculty from manual L+T+P
        total_tch = sum(r["tch"] for r in rows)
        min_wl = m["norm_hours"]
        deviation = total_tch - min_wl
        total_workload = total_tch + m["other_academic"]

        # Filter out the placeholder row for unassigned faculty
        subject_rows = [r for r in rows if r["course_code"]]

        result.append({
            "staff_id":     sid,
            "emp_code":     m["emp_code"],
            "faculty_name": m["faculty_name"],
            "designation":  m["designation"],
            "subjects":     subject_rows if subject_rows else rows,
            "total_tch":    total_tch,
            "min_workload": min_wl,
            "deviation":    deviation,
            "other_academic": m["other_academic"],
            "total_workload": total_workload,
            "remarks":      m["remarks"],
            "research_scholars": m["research_scholars"],
        })

    # FIX #6: sort faculty blocks by emp_code ASC
    result.sort(key=lambda b: b["emp_code"])

    return result


# ─── Excel generation ────────────────────────────────────────────────────────

def generate_master_workload_excel(
    academic_year: Optional[str] = None,
    semester_id: Optional[int] = None,
) -> bytes:
    """Generate the institutional master workload Excel sheet. Returns bytes."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Fetch and group data ──
    data, ay, sid = _fetch_workload_data(academic_year, semester_id)
    faculty_blocks = _group_by_faculty(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Workload"

    # ── Style definitions ──
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # FIX #8
    header_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    data_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)  # FIX #4
    num_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    last_col_letter = get_column_letter(NUM_COLS)

    # ── FIXED COLUMN WIDTHS (FIX #3) ──
    for col_num, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # ── HEADER ROWS (1-4) ──
    header_texts = [
        "HINDUSTAN INSTITUTE OF TECHNOLOGY AND SCIENCE",
        "SCHOOL OF BASIC AND APPLIED SCIENCES",
        "DEPARTMENT NAME: COMPUTER APPLICATIONS",
        f"MASTER WORKLOAD - SEMESTER {sid} {ay}",
    ]
    for row_idx, txt in enumerate(header_texts, start=1):
        ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
        cell = ws.cell(row=row_idx, column=1, value=txt)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── TABLE HEADER (row 5) ──
    TABLE_HEADER_ROW = 5
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=TABLE_HEADER_ROW, column=col_idx, value=col_name)
        cell.font = table_header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── DATA ROWS ──
    current_row = TABLE_HEADER_ROW + 1  # row 6
    serial_no = 0

    for block in faculty_blocks:
        serial_no += 1
        subjects = block["subjects"]
        num_rows = max(len(subjects), 1)
        start_row = current_row
        end_row = current_row + num_rows - 1

        for i, subj in enumerate(subjects):
            row = current_row + i
            has_course = bool(subj.get("course_code"))

            # Per-subject columns (one per row)
            # Col 5: UG/PG (center)
            ws.cell(row=row, column=5, value=subj["ug_pg"] if has_course else "").alignment = num_align
            # Col 6: Programme (left, wrap) — FIX #4
            ws.cell(row=row, column=6, value=subj["programme"] if has_course else "").alignment = left_wrap
            # Col 7: Course Category (center)
            ws.cell(row=row, column=7, value=subj["course_category"] if has_course else "").alignment = num_align
            # Col 8: SEM — FIX #5: convert to Roman
            sem_display = _to_roman(subj["semester"]) if has_course else ""
            ws.cell(row=row, column=8, value=sem_display).alignment = num_align
            # Col 9: Section — FIX #7: exact as stored
            ws.cell(row=row, column=9, value=subj["section"] if has_course else "").alignment = num_align
            # Col 10: Student Strength (center)
            ws.cell(row=row, column=10, value=subj["student_strength"] if has_course else "").alignment = num_align
            # Col 11: Course Code (left)
            ws.cell(row=row, column=11, value=subj["course_code"] if has_course else "").alignment = left_align
            # Col 12: Course Name (left, wrap) — FIX #4
            ws.cell(row=row, column=12, value=subj["course_name"] if has_course else "").alignment = left_wrap
            # Col 13: Complexity (center) — computed from TCH
            complexity = ""
            if has_course:
                tch_val = subj["tch"]
                if tch_val >= 5:
                    complexity = "TOUGH"
                elif tch_val in (2, 3):
                    complexity = "LOW"
                else:
                    complexity = "MEDIUM"
            ws.cell(row=row, column=13, value=complexity).alignment = num_align
            # Col 14: Credits (center)
            ws.cell(row=row, column=14, value=subj["credits"] if has_course else "").alignment = num_align
            # Col 15-17: L, T, P (center) — FIX #1: from raw values
            ws.cell(row=row, column=15, value=subj["l_assigned"] if has_course else "").alignment = num_align
            ws.cell(row=row, column=16, value=subj["t_assigned"] if has_course else "").alignment = num_align
            ws.cell(row=row, column=17, value=subj["p_assigned"] if has_course else "").alignment = num_align
            # Col 18: L+T+P (center) — FIX #1: computed manually
            ws.cell(row=row, column=18, value=subj["ltp"] if has_course else "").alignment = num_align
            # Col 19: TCH (center) — FIX #1: = L+T+P
            ws.cell(row=row, column=19, value=subj["tch"] if has_course else "").alignment = num_align

            # Apply border and font to all per-row cells
            for c in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                if cell.font == Font():
                    cell.font = data_font

        # ── Merged columns — faculty-level summary (FIX #2) ──
        # Write values ONCE into the START row only
        ws.cell(row=start_row, column=1, value=serial_no).alignment = num_align           # S No
        ws.cell(row=start_row, column=2, value=block["emp_code"]).alignment = left_align   # Emp Code
        ws.cell(row=start_row, column=3, value=block["faculty_name"]).alignment = left_align
        ws.cell(row=start_row, column=4, value=block["designation"]).alignment = left_align
        ws.cell(row=start_row, column=20, value=block["min_workload"]).alignment = num_align
        ws.cell(row=start_row, column=21, value=block["deviation"]).alignment = num_align
        # Col 22: Remarks (left, wrap) — FIX #4
        ws.cell(row=start_row, column=22, value=block["remarks"]).alignment = left_wrap
        ws.cell(row=start_row, column=23, value=block["other_academic"]).alignment = num_align
        ws.cell(row=start_row, column=24, value=block["total_workload"]).alignment = num_align
        # Col 25: No. of Research Scholars — FIX #9: ALWAYS BLANK
        ws.cell(row=start_row, column=25, value=None).alignment = num_align

        # Font + border for the summary cells
        for c in MERGE_COLS:
            col_num = c + 1
            cell = ws.cell(row=start_row, column=col_num)
            cell.font = data_font
            cell.border = thin_border

        # ── Vertical merge if multiple subject rows (FIX #2) ──
        if num_rows > 1:
            for c in MERGE_COLS:
                col_letter = get_column_letter(c + 1)
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")

        current_row = end_row + 1

    # ── Row heights ──
    for r in range(1, 5):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[TABLE_HEADER_ROW].height = 30

    # ── Freeze panes (freeze header + table header) ──
    ws.freeze_panes = f"A{TABLE_HEADER_ROW + 1}"

    # ── Save ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─── Snapshot-based Excel generation (NO DB queries) ─────────────────────────

def generate_from_snapshot(
    snapshot_data: list[dict],
    academic_year: str,
    semester_id: int,
) -> bytes:
    """
    Generate the institutional master workload Excel sheet from snapshot JSON.
    Zero database queries — uses only the pre-computed snapshot_data.

    Args:
        snapshot_data: list of faculty block dicts from workload_snapshot.snapshot_data
        academic_year: e.g. "2025-2026"
        semester_id: integer 1-6

    Returns:
        Excel file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Workload"

    # ── Style definitions ──
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    data_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    num_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    last_col_letter = get_column_letter(NUM_COLS)

    # ── Fixed column widths ──
    for col_num, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # ── Header rows (1-4) ──
    header_texts = [
        "HINDUSTAN INSTITUTE OF TECHNOLOGY AND SCIENCE",
        "SCHOOL OF BASIC AND APPLIED SCIENCES",
        "DEPARTMENT NAME: COMPUTER APPLICATIONS",
        f"MASTER WORKLOAD - SEMESTER {semester_id} {academic_year}",
    ]
    for row_idx, txt in enumerate(header_texts, start=1):
        ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
        cell = ws.cell(row=row_idx, column=1, value=txt)
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Table header (row 5) ──
    TABLE_HEADER_ROW = 5
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=TABLE_HEADER_ROW, column=col_idx, value=col_name)
        cell.font = table_header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Data rows from snapshot ──
    current_row = TABLE_HEADER_ROW + 1
    serial_no = 0

    for block in snapshot_data:
        serial_no += 1
        subjects = block.get("subjects", [])
        num_rows = max(len(subjects), 1)
        start_row = current_row
        end_row = current_row + num_rows - 1

        for i, subj in enumerate(subjects):
            row = current_row + i
            has_course = bool(subj.get("course_code"))

            ws.cell(row=row, column=5, value=subj.get("ug_pg", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=6, value=subj.get("programme", "") if has_course else "").alignment = left_wrap
            ws.cell(row=row, column=7, value=subj.get("course_category", "") if has_course else "").alignment = num_align
            sem_display = _to_roman(subj.get("semester", "")) if has_course else ""
            ws.cell(row=row, column=8, value=sem_display).alignment = num_align
            ws.cell(row=row, column=9, value=subj.get("section", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=10, value=subj.get("student_strength", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=11, value=subj.get("course_code", "") if has_course else "").alignment = left_align
            ws.cell(row=row, column=12, value=subj.get("course_name", "") if has_course else "").alignment = left_wrap
            # Col 13: Complexity (computed from TCH)
            complexity = ""
            if has_course:
                tch_val = subj.get("tch", 0)
                if tch_val >= 5:
                    complexity = "TOUGH"
                elif tch_val in (2, 3):
                    complexity = "LOW"
                else:
                    complexity = "MEDIUM"
            ws.cell(row=row, column=13, value=complexity).alignment = num_align
            ws.cell(row=row, column=14, value=subj.get("credits", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=15, value=subj.get("l", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=16, value=subj.get("t", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=17, value=subj.get("p", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=18, value=subj.get("ltp", "") if has_course else "").alignment = num_align
            ws.cell(row=row, column=19, value=subj.get("tch", "") if has_course else "").alignment = num_align

            for c in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                if cell.font == Font():
                    cell.font = data_font

        # ── Faculty-level merged columns ──
        ws.cell(row=start_row, column=1, value=serial_no).alignment = num_align
        ws.cell(row=start_row, column=2, value=block.get("emp_code", "")).alignment = left_align
        ws.cell(row=start_row, column=3, value=block.get("faculty_name", "")).alignment = left_align
        ws.cell(row=start_row, column=4, value=block.get("designation", "")).alignment = left_align
        ws.cell(row=start_row, column=20, value=block.get("min_workload", 12)).alignment = num_align
        ws.cell(row=start_row, column=21, value=block.get("deviation", 0)).alignment = num_align
        ws.cell(row=start_row, column=22, value=block.get("remarks", "")).alignment = left_wrap
        ws.cell(row=start_row, column=23, value=block.get("other_academic", 0)).alignment = num_align
        ws.cell(row=start_row, column=24, value=block.get("total_workload", 0)).alignment = num_align
        ws.cell(row=start_row, column=25, value=None).alignment = num_align  # Research Scholars: ALWAYS BLANK

        for c in MERGE_COLS:
            col_num = c + 1
            cell = ws.cell(row=start_row, column=col_num)
            cell.font = data_font
            cell.border = thin_border

        if num_rows > 1:
            for c in MERGE_COLS:
                col_letter = get_column_letter(c + 1)
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")

        current_row = end_row + 1

    # ── Row heights ──
    for r in range(1, 5):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[TABLE_HEADER_ROW].height = 30

    ws.freeze_panes = f"A{TABLE_HEADER_ROW + 1}"
    
    # ================================================================
    # SHEET 2: Workload (simplified view)
    # ================================================================
    ws2 = wb.create_sheet(title="workload")
    ws2.append(["Faculty Name", "Programme", "SEM", "Section", "Student Strength", "COURSE CODE", "COURSE NAME", "TCH"])
    
    for block in snapshot_data:
        faculty_name = block.get("faculty_name", "")
        for subj in block.get("subjects", []):
            if subj.get("course_code"):  # Only rows with actual courses
                ws2.append([
                    faculty_name,
                    subj.get("programme", ""),
                    _to_roman(subj.get("semester", "")),
                    subj.get("section", ""),
                    subj.get("student_strength", ""),
                    subj.get("course_code", ""),
                    subj.get("course_name", ""),
                    subj.get("tch", "")
                ])
    
    # ================================================================
    # SHEET 3: Faculty-list (just names)
    # ================================================================
    ws3 = wb.create_sheet(title="Faculty-list")
    ws3.append(["Faculty Name"])
    
    for block in snapshot_data:
        ws3.append([block.get("faculty_name", "")])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
